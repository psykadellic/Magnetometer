from openpyxl import load_workbook
from statistics import pvariance
from matplotlib import pyplot as plt
import os
import plotly.graph_objects as go
from classifier import classify, s_nothing, s_opened, s_closed


def showplt(col_x, col_y):
    plt.plot(col_x, col_y, linewidth=2, color='r')
    plt.title('Y-Direction of Microteslas')
    plt.xlabel('Time in Seconds')
    plt.ylabel('Microteslas')
    plt.show()
    #plt.savefig(file+'.png')

for file in os.listdir('random_xlsx_files'):
    wb = load_workbook(filename='random_xlsx_files/'+file)  # loops through each file
    ws = wb.active

    sheet_ranges = wb['Raw Data']
    data_length = len(sheet_ranges['A'])

    for col_x in ws.iter_cols(min_col=1, min_row=2, max_col=1, max_row=data_length, values_only=True):
        list(col_x)  # print(col_x) # x-axis, time in seconds

    for col_y in ws.iter_cols(min_col=3, min_row=2, max_col=3, max_row=data_length, values_only=True):
        list(col_y)  # y-axis, magnitude of y-direction microteslas

    showplt(col_x,col_y)
    print(file) #end function

fig = go.Figure(data=[go.Table(header=dict(values=['Opened', 'Closed', 'Nothing']),
                                   cells=dict(values=[list(s_opened()), list(s_closed()), list(s_nothing())]))
                          ])

'''
fig.add_layout_image(
        dict(
            source=file+".png",
            xref="x",
            yref="y",
            x=0,
            y=0,
            sizex=100,
            sizey=100,
            layer="below")
)
'''
fig.show()