from openpyxl import load_workbook
from statistics import pvariance
from matplotlib import pyplot as plt
import os, random


def showplt(col_x, col_y):      #all_files.py
    plt.plot(col_x, col_y, linewidth=2, color='r')
    plt.title('Y-Direction of Microteslas')
    plt.xlabel('Time in Seconds')
    plt.ylabel('Microteslas')
    plt.show()
    #plt.savefig(file+'.png')

def classify():
    stack_classify = []  # confusion_matrix.py

    stack_opened = []  # display for allfiles.py
    stack_closed = []  # allfiles.py
    stack_nothing = []  # allfiles.py

    for file in os.listdir('random_xlsx_files'):
        wb = load_workbook(filename='random_xlsx_files/' + file)  # loop through each file
        ws = wb.active

        sheet_ranges = wb['Raw Data']
        data_length = len(sheet_ranges['A'])

        for col_x in ws.iter_cols(min_col=1, min_row=2, max_col=1, max_row=data_length, values_only=True):
            list(col_x)  # print(col_x) # x-axis, time in seconds

        for col_y in ws.iter_cols(min_col=3, min_row=2, max_col=3, max_row=data_length, values_only=True):
            list(col_y)  # y-axis, magnitude of y-direction microteslas

        #showplt(col_x, col_y)  # allfiles.py

        nothing = 'Nothing'
        opened = 'Close to open'
        closed = 'Open to close'

        if pvariance(col_y) < 0.5:
            classification = nothing
            # print(classification)
            stack_nothing.append(file)
            stack_classify.append(classification)
        elif pvariance(col_y) >= 0.5:
            if sheet_ranges['C2'].value > sheet_ranges['C' + str(data_length)].value:
                classification = opened
                # print(classification)
                stack_opened.append(file)
                stack_classify.append(classification)
            else:
                classification = closed
                # print(classification)
                stack_closed.append(file)
                stack_classify.append(classification)
    return stack_classify, stack_nothing, stack_opened, stack_closed

stack_classify2, stack_nothing2, stack_opened2, stack_closed2 = classify()
#print(stack_classify2)

def s_classify():   # confusion_matrix.py
    return stack_classify2

def s_nothing():    # all_files.py
    return stack_nothing2

def s_opened():     # all_files.py
    return stack_opened2

def s_closed():     # all_files.py
    return stack_closed2

def workb_ran ():
    sample = random.os.listdir('random_xlsx_files')
    wb = load_workbook(filename='random_xlsx_files/' + sample)  # loop through each file
    ws = wb.active

    sheet_ranges = wb['Raw Data']
    data_length = len(sheet_ranges['A'])

    for col_x in ws.iter_cols(min_col=1, min_row=2, max_col=1, max_row=data_length, values_only=True):
        list(col_x)  # print(col_x) # x-axis, time in seconds

    for col_y in ws.iter_cols(min_col=3, min_row=2, max_col=3, max_row=data_length, values_only=True):
        list(col_y)  # y-axis, magnitude of y-direction microteslas



